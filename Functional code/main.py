import csv
import json
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill

dic_id_map = {}
dic_parent_map = {}
dic_assignee = {}
dic_story_map = {}
new_stories = []

with open('conf.json', encoding='utf-8') as config_file:
    conf_data = json.load(config_file)

def list_stories(key, id, parent, row):
    if parent == '':
        dic_story_map[key] = row
        map_parent(key, id)
    else:
        story_key = dic_id_map[parent]
        map_parent(story_key, id)


def map_parent(key, id):
    if key not in dic_parent_map.keys():
        dic_parent_map[key] = []
    dic_parent_map[key].append(id)

def map_columns(array, name):
    for i in range(0, len(array)):
        if array[i] == name:
            return i
            break

def write_row(row_index, issue_key):
    write_assignee(row_index, issue_key)
    sheet.cell(row=row_index, column=2).value = dic_story_map[issue_key][status_col]
    sheet.cell(row=row_index, column=4).value = dic_story_map[issue_key][priority_col]
    sheet.cell(row=row_index, column=col_max).value = dic_story_map[issue_key][summary_col]
    if dic_story_map[issue_key][story_point_col] =='':
        sheet.cell(row=row_index, column=5).value = 0
    else:
        sheet.cell(row=row_index, column=5).value = float(dic_story_map[issue_key][story_point_col])


def write_assignee(row_index,issue_key):
    for j in range(6, col_max):
        assignee = sheet.cell(row=1, column=j).value
        assignee = conf_data[assignee]
        is_assignee = test_assignee(assignee, issue_key)
        if is_assignee == True:
            sheet.cell(row=row_index, column=j).value = 'x'

def test_assignee(assignee, issue_key):
    issue_id = dic_story_map[issue_key][issue_id_col]
    is_assignee = None
    if dic_parent_map != {}:
        parent_id = dic_parent_map[issue_key]
        for item in dic_assignee[assignee]:
            if item == issue_id:
                is_assignee = True
            elif item in parent_id:
                is_assignee = True
    else:
        for item in dic_assignee[assignee]:
            if item == issue_id:
                is_assignee = True
    return is_assignee


# file name JIRA.csv, separators ,
with open('JIRA.csv') as csvfile:
    readCSV = csv.reader(csvfile, delimiter=',')
    data = list(readCSV)
    row_id = 1
    #maps each header to a column number based on the label
    for row in data:
        if row_id == 1:
            header = row
            issue_key_col = map_columns(row, 'Issue key')
            issue_id_col = map_columns(row, 'Issue id')
            parent_id_col = map_columns(row,'Parent id')
            summary_col = map_columns(row,'Summary')
            assignee_col = map_columns(row, 'Assignee')
            priority_col = map_columns(row, 'Priority')
            status_col = map_columns(row, 'Status')
            story_point_col = map_columns(row, 'Custom field (Story Points)')
            row_id += 1
        else:
            #associates each issue id to a key
            dic_id_map[row[issue_id_col]] = row[issue_key_col]
            if row[assignee_col] not in dic_assignee.keys():
                dic_assignee[row[assignee_col]] = []
            dic_assignee[row[assignee_col]].append(row[issue_id_col])

    #build a list of the stories
    for row in data:
        if row == header:
            continue
        if parent_id_col != None:
            list_stories(row[issue_key_col], row[issue_id_col], row[parent_id_col], row)
        else:
            dic_story_map[row[issue_key_col]] = row
        row_id += 1

#generates the output file
output_file = load_workbook('Sprint_matrix.xlsx')
sheet = output_file.active
row_max = sheet.max_row
col_max = sheet.max_column

#checks for existing lines and updates them (specifically useful for update)
for key in dic_story_map:
    for i in range(3, row_max):
        c = i
        issue_key = sheet.cell(row=i, column=3).value
        if issue_key != None:
            if issue_key == key:
                write_row(i, issue_key)
                break
            else:
                continue
    if c == row_max - 1:
        if key not in new_stories:
            new_stories.append(key)

#adds new lines in the end of the file (specifically useful for first creation)
index = row_max
for item in new_stories:
    if dic_story_map[item][status_col] == 'Done':
        sheet.insert_rows(4)
        sheet.cell(row=4, column=3).value = item
        write_row(4, item)
    else:
        write_row(index, item)
        sheet.cell(row=index, column=3).value = item

    index += 1

#computes the total sums
total_row = {}
row_max = sheet.max_row
for i in range(3, row_max):
    category = sheet.cell(row=i, column=1).value
    if category != None and category != '':
        total_row[category] = []
        total_row[category].append(i)

for key in total_row.keys():
    for i in range(total_row[key][0], row_max):
        total = sheet.cell(row=i, column=4).value
        if total == 'Total':
            total_row[key].append(i)
            break

for key in total_row.keys():
    if key != None and key != 'To be sorted':
        top_cell = 'E' + str(total_row[key][0])
        total_cell = 'E' + str(total_row[key][1] - 1)
        total = sheet.cell(row=total_row[key][1], column=5)
        total.value = '= SUM(' + top_cell + ':' + total_cell + ')'

#highlight the removed stories
cell_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type = "solid")
for i in range(3, row_max):
    story_key = sheet.cell(row=i, column=3)
    if story_key.value != None:
        if story_key.value not in dic_story_map.keys():
            for j in range(1, col_max +1):
                sheet.cell(row=i, column=j).fill = cell_fill

output_file.save('Sprint_matrix_update.xlsx')